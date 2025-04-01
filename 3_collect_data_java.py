import os
from tree_sitter import Node, Tree
from tree_sitter_languages import get_parser
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ERRORS = set(['AclNotFoundException', 'ActivationException', 'UnknownGroupException', 'UnknownObjectException', 'AlreadyBoundException', 'ApplicationException', 'AWTException', 'BackingStoreException', 'BadAttributeValueExpException', 'BadBinaryOpValueExpException', 'BadLocationException', 'BadStringOperationException', 'BrokenBarrierException', 'CertificateException', 'CertificateEncodingException', 'CertificateExpiredException', 'CertificateNotYetValidException', 'CertificateParsingException', 'CloneNotSupportedException', 'ServerCloneException', 'DataFormatException', 'DatatypeConfigurationException', 'DestroyFailedException', 'ExecutionException', 'ExpandVetoException', 'FontFormatException', 'GeneralSecurityException', 'BadPaddingException', 'AEADBadTagException', 'CertificateException', 'CertificateEncodingException', 'CertificateExpiredException', 'CertificateNotYetValidException', 'CertificateParsingException', 'CertificateRevokedException', 'CertPathBuilderException', 'CertPathValidatorException', 'CertStoreException', 'CRLException', 'DigestException', 'ExemptionMechanismException', 'IllegalBlockSizeException', 'InvalidAlgorithmParameterException', 'InvalidKeySpecException', 'InvalidParameterSpecException', 'KeyException', 'InvalidKeyException', 'KeyManagementException', 'KeyStoreException', 'LoginException', 'AccountException', 'AccountExpiredException', 'AccountLockedException', 'AccountNotFoundException', 'CredentialException', 'CredentialExpiredException', 'CredentialNotFoundException', 'FailedLoginException', 'NoSuchAlgorithmException', 'NoSuchPaddingException', 'NoSuchProviderException', 'ShortBufferException', 'SignatureException', 'UnrecoverableEntryException', 'UnrecoverableKeyException', 'GSSException', 'IllegalClassFormatException', 'InterruptedException', 'IntrospectionException', 'InvalidApplicationException', 'InvalidMidiDataException', 'InvalidPreferencesFormatException', 'InvalidTargetObjectTypeException', 'IOException', 'ChangedCharSetException', 'CharacterCodingException', 'MalformedInputException', 'UnmappableCharacterException', 'CharConversionException', 'ClosedChannelException', 'AsynchronousCloseException', 'ClosedByInterruptException', 'EOFException', 'FileLockInterruptionException', 'FileNotFoundException', 'FilerException', 'FileSystemException', 'AccessDeniedException', 'AtomicMoveNotSupportedException', 'DirectoryNotEmptyException', 'FileAlreadyExistsException', 'FileSystemLoopException', 'NoSuchFileException', 'NotDirectoryException', 'NotLinkException', 'HttpRetryException', 'IIOException', 'IIOInvalidTreeException', 'InterruptedByTimeoutException', 'InterruptedIOException', 'SocketTimeoutException', 'InvalidPropertiesFormatException', 'JMXProviderException', 'JMXServerErrorException', 'MalformedURLException', 'ObjectStreamException', 'InvalidClassException', 'InvalidObjectException', 'NotActiveException', 'NotSerializableException', 'OptionalDataException', 'StreamCorruptedException', 'WriteAbortedException', 'ProtocolException', 'RemoteException', 'AccessException', 'ActivateFailedException', 'ActivityCompletedException', 'ActivityRequiredException', 'ConnectException', 'ConnectIOException', 'ExportException', 'SocketSecurityException', 'InvalidActivityException', 'InvalidTransactionException', 'MarshalException', 'NoSuchObjectException', 'ServerError', 'ServerException', 'ServerRuntimeException', 'SkeletonMismatchException', 'SkeletonNotFoundException', 'StubNotFoundException', 'TransactionRequiredException', 'TransactionRolledbackException', 'UnexpectedException', 'UnknownHostException', 'UnmarshalException', 'SaslException', 'AuthenticationException', 'SocketException', 'BindException', 'ConnectException', 'NoRouteToHostException', 'PortUnreachableException', 'SSLException', 'SSLHandshakeException', 'SSLKeyException', 'SSLPeerUnverifiedException', 'SSLProtocolException', 'SyncFailedException', 'UnknownHostException', 'UnknownServiceException', 'UnsupportedDataTypeException', 'UnsupportedEncodingException', 'UserPrincipalNotFoundException', 'UTFDataFormatException', 'ZipException', 'JarException', 'JAXBException', 'MarshalException', 'PropertyException', 'UnmarshalException', 'ValidationException', 'JMException', 'MBeanException', 'MBeanRegistrationException', 'OpenDataException', 'OperationsException', 'AttributeNotFoundException', 'InstanceAlreadyExistsException', 'InstanceNotFoundException', 'IntrospectionException', 'InvalidAttributeValueException', 'ListenerNotFoundException', 'MalformedObjectNameException', 'NotCompliantMBeanException', 'ServiceNotFoundException', 'ReflectionException', 'RelationException', 'InvalidRelationIdException', 'InvalidRelationServiceException', 'InvalidRelationTypeException', 'InvalidRoleInfoException', 'InvalidRoleValueException', 'RelationNotFoundException', 'RelationServiceNotRegisteredException', 'RelationTypeNotFoundException', 'RoleInfoNotFoundException', 'RoleNotFoundException', 'KeySelectorException', 'LambdaConversionException', 'LastOwnerException', 'LineUnavailableException', 'MarshalException', 'MidiUnavailableException', 'MimeTypeParseException', 'MimeTypeParseException', 'NamingException', 'AttributeInUseException', 'AttributeModificationException', 'CannotProceedException', 'CommunicationException', 'ConfigurationException', 'ContextNotEmptyException', 'InsufficientResourcesException', 'InterruptedNamingException', 'InvalidAttributeIdentifierException', 'InvalidAttributesException', 'InvalidAttributeValueException', 'InvalidNameException', 'InvalidSearchControlsException', 'InvalidSearchFilterException', 'LimitExceededException', 'SizeLimitExceededException', 'TimeLimitExceededException', 'LinkException', 'LinkLoopException', 'MalformedLinkException', 'NameAlreadyBoundException', 'NameNotFoundException', 'NamingSecurityException', 'AuthenticationException', 'AuthenticationNotSupportedException', 'NoPermissionException', 'NoInitialContextException', 'NoSuchAttributeException', 'NotContextException', 'OperationNotSupportedException', 'PartialResultException', 'ReferralException', 'LdapReferralException', 'SchemaViolationException', 'ServiceUnavailableException', 'NoninvertibleTransformException', 'NotBoundException', 'NotOwnerException', 'ParseException', 'ParserConfigurationException', 'PrinterException', 'PrinterAbortException', 'PrinterIOException', 'PrintException', 'PrivilegedActionException', 'PropertyVetoException', 'ReflectiveOperationException', 'ClassNotFoundException', 'IllegalAccessException', 'InstantiationException', 'InvocationTargetException', 'NoSuchFieldException', 'NoSuchMethodException', 'RefreshFailedException', 'RemarshalException', 'RuntimeException', 'AnnotationTypeMismatchException', 'ArithmeticException', 'ArrayStoreException', 'BufferOverflowException', 'BufferUnderflowException', 'CannotRedoException', 'CannotUndoException', 'ClassCastException', 'CMMException', 'CompletionException', 'ConcurrentModificationException', 'DirectoryIteratorException', 'DataBindingException', 'DateTimeException', 'DateTimeParseException', 'UnsupportedTemporalTypeException', 'ZoneRulesException', 'DOMException', 'EmptyStackException', 'EnumConstantNotPresentException', 'EventException', 'FileSystemAlreadyExistsException', 'FileSystemNotFoundException', 'IllegalArgumentException', 'IllegalChannelGroupException', 'IllegalCharsetNameException', 'IllegalFormatException', 'DuplicateFormatFlagsException', 'FormatFlagsConversionMismatchException', 'IllegalFormatCodePointException', 'IllegalFormatConversionException', 'IllegalFormatFlagsException', 'IllegalFormatPrecisionException', 'IllegalFormatWidthException', 'MissingFormatArgumentException', 'MissingFormatWidthException', 'UnknownFormatConversionException', 'UnknownFormatFlagsException', 'IllegalSelectorException', 'IllegalThreadStateException', 'InvalidKeyException', 'InvalidOpenTypeException', 'InvalidParameterException', 'InvalidPathException', 'KeyAlreadyExistsException', 'NumberFormatException', 'PatternSyntaxException', 'ProviderMismatchException', 'UnresolvedAddressException', 'UnsupportedAddressTypeException', 'UnsupportedCharsetException', 'IllegalMonitorStateException', 'IllegalPathStateException', 'IllegalStateException', 'AcceptPendingException', 'AlreadyBoundException', 'AlreadyConnectedException', 'CancellationException', 'CancelledKeyException', 'ClosedDirectoryStreamException', 'ClosedFileSystemException', 'ClosedSelectorException', 'ClosedWatchServiceException', 'ConnectionPendingException', 'FormatterClosedException', 'IllegalBlockingModeException', 'IllegalComponentStateException', 'InvalidDnDOperationException', 'InvalidMarkException', 'NoConnectionPendingException', 'NonReadableChannelException', 'NonWritableChannelException', 'NotYetBoundException', 'NotYetConnectedException', 'OverlappingFileLockException', 'ReadPendingException', 'ShutdownChannelGroupException', 'WritePendingException', 'IllformedLocaleException', 'ImagingOpException', 'IncompleteAnnotationException', 'IndexOutOfBoundsException', 'ArrayIndexOutOfBoundsException', 'StringIndexOutOfBoundsException', 'JMRuntimeException', 'MonitorSettingException', 'RuntimeErrorException', 'RuntimeMBeanException', 'RuntimeOperationsException', 'LSException', 'MalformedParameterizedTypeException', 'MalformedParametersException', 'MirroredTypesException', 'MirroredTypeException', 'MissingResourceException', 'NegativeArraySizeException', 'NoSuchElementException', 'InputMismatchException', 'NoSuchMechanismException', 'NullPointerException', 'ProfileDataException', 'ProviderException', 'ProviderNotFoundException', 'RasterFormatException', 'RejectedExecutionException', 'SecurityException', 'AccessControlException', 'RMISecurityException', 'SystemException', 'ACTIVITY_COMPLETED', 'ACTIVITY_REQUIRED', 'BAD_CONTEXT', 'BAD_INV_ORDER', 'BAD_OPERATION', 'BAD_PARAM', 'BAD_QOS', 'BAD_TYPECODE', 'CODESET_INCOMPATIBLE', 'COMM_FAILURE', 'DATA_CONVERSION', 'FREE_MEM', 'IMP_LIMIT', 'IndirectionException', 'INITIALIZE', 'INTERNAL', 'INTF_REPOS', 'INV_FLAG', 'INV_IDENT', 'INV_OBJREF', 'INV_POLICY', 'INVALID_ACTIVITY', 'INVALID_TRANSACTION', 'MARSHAL', 'NO_IMPLEMENT', 'NO_MEMORY', 'NO_PERMISSION', 'NO_RESOURCES', 'NO_RESPONSE', 'OBJ_ADAPTER', 'OBJECT_NOT_EXIST', 'PERSIST_STORE', 'REBIND', 'TIMEOUT', 'TRANSACTION_MODE', 'TRANSACTION_REQUIRED', 'TRANSACTION_ROLLEDBACK', 'TRANSACTION_UNAVAILABLE', 'TRANSIENT', 'UNKNOWN', 'UnknownException', 'TypeConstraintException', 'TypeNotPresentException', 'UncheckedIOException', 'UndeclaredThrowableException', 'UnknownEntityException', 'UnknownAnnotationValueException', 'UnknownElementException', 'UnknownTypeException', 'UnmodifiableSetException', 'UnsupportedOperationException', 'HeadlessException', 'ReadOnlyBufferException', 'ReadOnlyFileSystemException', 'WebServiceException', 'ProtocolException', 'HTTPException', 'SOAPFaultException', 'WrongMethodTypeException', 'SAXException', 'SAXNotRecognizedException', 'SAXNotSupportedException', 'SAXParseException', 'ScriptException', 'ServerNotActiveException', 'SOAPException', 'SQLException', 'BatchUpdateException', 'RowSetWarning', 'SerialException', 'SQLClientInfoException', 'SQLNonTransientException', 'SQLDataException', 'SQLFeatureNotSupportedException', 'SQLIntegrityConstraintViolationException', 'SQLInvalidAuthorizationSpecException', 'SQLNonTransientConnectionException', 'SQLSyntaxErrorException', 'SQLRecoverableException', 'SQLTransientException', 'SQLTimeoutException', 'SQLTransactionRollbackException', 'SQLTransientConnectionException', 'SQLWarning', 'DataTruncation', 'SyncFactoryException', 'SyncProviderException', 'TimeoutException', 'TooManyListenersException', 'TransformerException', 'TransformerConfigurationException', 'TransformException', 'UnmodifiableClassException', 'UnsupportedAudioFileException', 'UnsupportedCallbackException', 'UnsupportedFlavorException', 'UnsupportedLookAndFeelException', 'URIReferenceException', 'URISyntaxException', 'UserException', 'AdapterAlreadyExists', 'AdapterInactive', 'AdapterNonExistent', 'AlreadyBound', 'BadKind', 'Bounds', 'Bounds', 'CannotProceed', 'DuplicateName', 'FormatMismatch', 'ForwardRequest', 'ForwardRequest', 'InconsistentTypeCode', 'InconsistentTypeCode', 'Invalid', 'InvalidAddress', 'InvalidName', 'InvalidName', 'InvalidName', 'InvalidPolicy', 'InvalidSeq', 'InvalidSlot', 'InvalidTypeForEncoding', 'InvalidValue', 'InvalidValue', 'NoContext', 'NoServant', 'NotEmpty', 'NotFound', 'ObjectAlreadyActive', 'ObjectNotActive', 'PolicyError', 'ServantAlreadyActive', 'ServantNotActive', 'TypeMismatch', 'TypeMismatch', 'TypeMismatch', 'UnknownEncoding', 'UnknownUserException', 'WrongAdapter', 'WrongPolicy', 'WrongTransaction', 'XAException', 'XMLParseException', 'XMLSignatureException', 'XMLStreamException', 'XPathException', 'XPathExpressionException', 'XPathFunctionException', 'XPathFactoryConfigurationException'])

parser = get_parser('java')

# https://github.com/tree-sitter/py-tree-sitter/issues/33
def traverse_tree(tree: Tree):
	cursor = tree.walk()

	reached_root = False
	while reached_root == False:
		yield cursor.node

		if cursor.goto_first_child():
			continue

		if cursor.goto_next_sibling():
			continue

		retracing = True
		while retracing:
			if not cursor.goto_parent():
				retracing = False
				reached_root = True

			if cursor.goto_next_sibling():
				retracing = False

def parse_integer_literal(node: Node) -> int:
	return int(node.text.decode().replace('L', '').replace('l', '').replace('_', ''))

# 1
class_member_type_counts = defaultdict(int)
parameter_type_counts = defaultdict(int)
local_var_type_counts = defaultdict(int)

# 2
method_parameter_counts = defaultdict(int)

# 3
class_line_counts = defaultdict(int)
method_line_counts = defaultdict(int)

# 4
class_method_counts = defaultdict(int)

# 5
class_member_counts = defaultdict(int)

# 6
permission_default_counts = defaultdict(int)
permission_private_counts = defaultdict(int)
permission_public_counts = defaultdict(int)
permission_protected_counts = defaultdict(int)

# 7
class_uses_inheritance = 0
class_uses_interfaces = 0
class_uses_both = 0
class_uses_neither = 0

# 8
recursive_method_call_counts = defaultdict(int)
nonrecursive_method_call_counts = defaultdict(int)
total_method_count = 0

# 9
does_not_throw = 0
throws_generic_counts = defaultdict(int)
throws_jre_counts = defaultdict(int)
throws_custom_counts = defaultdict(int)
throws_counts = defaultdict(int)

# 10
does_not_catch = 0
catches_generic_counts = defaultdict(int)
catches_jre_counts = defaultdict(int)
catches_custom_counts = defaultdict(int)

# 11
if_block_counts = defaultdict(int)
while_block_counts = defaultdict(int)
for_block_counts = defaultdict(int)
enhanced_for_block_counts = defaultdict(int)
switch_block_counts = defaultdict(int)

# 12
defined_final_constants = defaultdict(int)
used_constant_numbers = defaultdict(int)
uses_constant_in_and_out_range_count = 0
uses_constant_in_range_count = 0
uses_constant_out_range_count = 0

say_file = False

def handle_node(node: Node, constants: dict[str, int]):
	global say_file
	global class_uses_both, class_uses_inheritance, class_uses_interfaces, class_uses_neither, total_method_count, does_not_throw, does_not_catch, uses_constant_in_and_out_range_count, uses_constant_in_range_count, uses_constant_out_range_count

	# 1
	if node.type == 'local_variable_declaration':
		local_var_type_counts[node.child_by_field_name('type').text.decode()] += 1

	# 1, 2
	if node.type == 'formal_parameters':
		cur = node.walk()
		method_parameter_counts[cur.node.named_child_count] += 1

		cur.goto_first_child()
		while True:
			if cur.node.type == 'formal_parameter':
				parameter_type_counts[cur.node.child_by_field_name('type').text.decode()] += 1
			if not cur.goto_next_sibling():
				break

	# 3
	if node.type == 'class_declaration':
		lc = node.text.count(b'\n')
		class_line_counts[lc] += 1
	elif node.type == 'method_declaration':
		lc = node.text.count(b'\n')
		method_line_counts[lc] += 1
		# if lc == 0:
		# 	say_file = True

	# 4, 5, 6, 12
	if node.type == 'class_body':
		cur = node.walk()
		cur.goto_first_child()

		# 4
		method_count = 0

		# 5
		member_count = 0

		# 6
		permission_default_count = 0
		permission_private_count = 0
		permission_public_count = 0
		permission_protected_count = 0

		while cur.goto_next_sibling():
			# 4
			if cur.node.type == 'method_declaration':
				method_count += 1

			# 1, 5, 6, 12
			if cur.node.type == 'field_declaration':
				# 1
				class_member_type_counts[cur.node.child_by_field_name('type').text.decode()] += 1

				# 5
				member_count += 1

				# 6
				perms = cur.node.child(0).text
				if b'private' in perms:
					permission_private_count += 1
				elif b'public' in perms:
					permission_public_count += 1
				elif b'protected' in perms:
					permission_protected_count += 1
				else:
					permission_default_count += 1

				# 12
				if b'final' in perms:
					v = cur.node.child_by_field_name('declarator').child_by_field_name('value')
					if v and v.type == 'decimal_integer_literal':
						k = cur.node.child_by_field_name('declarator').child_by_field_name('name').text
						n = parse_integer_literal(v)
						constants[k.decode()] = n
						defined_final_constants[n] += 1

		class_method_counts[method_count] += 1
		class_member_counts[member_count] += 1

		permission_default_counts[permission_default_count] += 1
		permission_private_counts[permission_private_count] += 1
		permission_public_counts[permission_public_count] += 1
		permission_protected_counts[permission_protected_count] += 1

	# 7
	if node.type == 'class_declaration':
		cur = node.walk()
		cur.goto_first_child()

		uses_inheritance = False
		uses_interfaces = False

		while cur.goto_next_sibling():
			if cur.node.type == 'superclass':
				uses_inheritance = True

			if cur.node.type == 'super_interfaces':
				uses_interfaces = True

		class_uses_both += uses_inheritance and uses_interfaces
		class_uses_inheritance += uses_inheritance
		class_uses_interfaces += uses_interfaces
		class_uses_neither += not uses_inheritance and not uses_interfaces

	# 8, 9, 10, 11, 12
	if node.type == 'method_declaration':
		# 9
		throws = node.child_by_field_name('parameters').next_sibling
		if throws.type == 'throws':
			cur = throws.walk()
			cur.goto_first_child()

			throws_generic_count = 0
			throws_jre_count = 0
			throws_custom_count = 0

			while cur.goto_next_sibling():
				if cur.node.type == 'type_identifier':
					err_name = cur.node.text.decode()
					if err_name == 'Exception':
						throws_generic_count += 1
					elif err_name in ERRORS:
						throws_jre_count += 1
					else:
						throws_custom_count += 1

			throws_generic_counts[throws_generic_count] += 1
			throws_jre_counts[throws_jre_count] += 1
			throws_custom_counts[throws_custom_count] += 1
			throws_counts[throws_generic_count + throws_jre_count + throws_custom_count] += 1
		else:
			does_not_throw += 1

		body = node.child_by_field_name('body')
		if body is not None:
			recursive_call_count = 0
			nonrecursive_call_count = 0

			declared_method = node.child_by_field_name('name').text

			if_block_count = 0
			while_block_count = 0
			for_block_count = 0
			enhanced_for_block_count = 0
			switch_block_count = 0

			catches_generic_count = 0
			catches_jre_count = 0
			catches_custom_count = 0

			# 12
			uses_constant_in_range = False
			uses_constant_out_range = False

			for fn_node in traverse_tree(body):
				if fn_node.type == 'method_invocation':
					called_method = fn_node.child_by_field_name('name').text
					obj = fn_node.child_by_field_name('object')

					if called_method == declared_method and (obj is None or obj.text.decode().startswith('this')):
						recursive_call_count += 1
					else:
						nonrecursive_call_count += 1

				if fn_node.type == 'if_statement':
					if_block_count += 1

				if fn_node.type == 'while_statement':
					while_block_count += 1

				if fn_node.type == 'for_statement':
					for_block_count += 1

				if fn_node.type == 'enhanced_for_statement':
					enhanced_for_block_count += 1

				if fn_node.type == 'switch_expression':
					switch_block_count += 1
				
				if fn_node.type == 'catch':
					cfp = fn_node.next_sibling.next_sibling
					if cfp.type == 'catch_formal_parameter':
						c = cfp.walk()
						c.goto_first_child()
						c.goto_first_child()
						while True:
							if c.node.type != '|':
								err_name = c.node.text.decode()
								if err_name == 'Exception':
									catches_generic_count += 1
								elif err_name in ERRORS:
									catches_jre_count += 1
								else:
									catches_custom_count += 1

							if not c.goto_next_sibling():
								break

				# 12
				num = None
				if fn_node.type == 'decimal_integer_literal':
					num = parse_integer_literal(fn_node)
				if fn_node.type == 'name' and fn_node.text.decode() in constants:
					num = constants[fn_node.text.decode()]
				if num:
					if abs(num) <= 32:
						uses_constant_in_range = True
					else:
						uses_constant_out_range = True

			# 8
			recursive_method_call_counts[recursive_call_count] += 1
			nonrecursive_method_call_counts[nonrecursive_call_count] += 1
			total_method_count += 1

			# 10
			if catches_generic_count == 0 and catches_jre_count == 0 and catches_custom_count == 0:
				does_not_catch += 1

			catches_generic_counts[catches_generic_count] += 1
			catches_jre_counts[catches_jre_count] += 1
			catches_custom_counts[catches_custom_count] += 1

			# 11
			if_block_counts[if_block_count] += 1
			while_block_counts[while_block_count] += 1
			for_block_counts[for_block_count] += 1
			enhanced_for_block_counts[enhanced_for_block_count] += 1
			switch_block_counts[switch_block_count] += 1

			# 12
			if uses_constant_in_range and uses_constant_out_range:
				uses_constant_in_and_out_range_count += 1
			else:
				uses_constant_in_range_count += uses_constant_in_range
				uses_constant_out_range_count += uses_constant_out_range

	# 12
	if node.type == 'decimal_integer_literal':
		# avoid final constant declarations from this count
		if b'final' not in node.parent.parent.text:
			used_constant_numbers[parse_integer_literal(node)] += 1

	# 12
	if node.type == 'name' and node.text.decode() in constants:
		used_constant_numbers[constants[node.text.decode()]] += 1

i = 0

for root, dirs, files in os.walk('repos\\Java'):
	# if i > 20000:
	# 	break

	for file in files:
		if file.endswith('.java'):
			i += 1

			if i % 10000 == 0:
				print(i)

			try:
				with open(os.path.join(root, file), 'rb') as f:
					constants = {}
					# if i < 313:
					# 	continue

					src = f.read()
					tree = parser.parse(src)

					# if not (b': ' in src and b'for (' in src):
					# 	continue

					for node in traverse_tree(tree):
						handle_node(node, constants)

					if say_file:
						say_file = False
						print(os.path.join(root, file), 'rb')
					# print(src.decode())
					# print(tree.root_node.sexp())
					# exit()
					# exit()
			except:
				continue

wb = Workbook()

def save_dict(name, data: dict):
	ws: Worksheet = wb.create_sheet(name[:31])
	ws.cell(1, 1, 'Value')
	ws.cell(1, 2, 'Quantity')
	ws.cell(1, 6, name)
	for i, (k, v) in enumerate(data.items()):
		# print(i, k, v)
		ws.cell(i + 2, 1, k)
		ws.cell(i + 2, 2, v)

# print('(1)', class_member_type_counts, parameter_type_counts, local_var_type_counts)
# print('(2)', method_parameter_counts)
# print('(3)', class_line_counts, method_line_counts)
# print('(4)', class_method_counts)
# print('(5)', class_member_counts)
# print('(6)', permission_default_counts, permission_public_counts, permission_private_counts, permission_protected_counts)
# print('(7)', class_uses_both, class_uses_inheritance, class_uses_interfaces, class_uses_neither)
# print('(8)', recursive_method_call_counts, nonrecursive_method_call_counts, total_method_count)
# print('(9)', does_not_throw, throws_generic_counts, throws_jre_counts, throws_custom_counts, throws_counts)
# print('(10)', does_not_catch, catches_generic_counts, catches_jre_counts, catches_custom_counts)
# print('(11)', if_block_counts, while_block_counts, for_block_counts, enhanced_for_block_counts, switch_block_counts)
# print('(12)', uses_constant_in_and_out_range_count, uses_constant_in_range_count, uses_constant_out_range_count, defined_final_constants, used_constant_numbers)
# print('scanned', i, 'files')

save_dict('Total numbers', {
	'total number of analyzed java files': i + 1,
	'total number of analyzed methods': total_method_count,
	'total number of analyzed classes': class_uses_neither + class_uses_inheritance + class_uses_interfaces - class_uses_both,
})

save_dict('Q1 Class Member Type Counts', class_member_type_counts)
save_dict('Q1 Function Parameter Type Counts', parameter_type_counts)
save_dict('Q1 Local Variable Type Counts', local_var_type_counts)
save_dict('Q2 Method Parameter Counts', method_parameter_counts)
save_dict('Q3 Class Line Counts', class_line_counts)
save_dict('Q3 Method Line Counts', method_line_counts)
save_dict('Q4 Class Method Counts', class_method_counts)
save_dict('Q5 Class Member Counts', class_member_counts)
save_dict('Q6 Class Member Default Permission Count', permission_default_counts)
save_dict('Q6 Class Member Public Count', permission_public_counts)
save_dict('Q6 Class Member Private Count', permission_private_counts)
save_dict('Q6 Class Member Protected Count', permission_protected_counts)
save_dict('Q7 Class Uses Inheritance Interfaces', {
	'Classes using inheritance & interface': class_uses_both,
	'Classes using inheritance': class_uses_inheritance,
	'Classes using interface': class_uses_interfaces,
	'Classes using neither inheritance nor interface': class_uses_neither,
})
save_dict('Q8 Number of recursive method calls in methods', recursive_method_call_counts)
save_dict('Q8 Number of non-recursive method calls in methods', nonrecursive_method_call_counts)
save_dict('Q9 Number of generic errors thrown in methods ', throws_generic_counts)
save_dict('Q9 Number of JRE errors thrown in methods ', throws_jre_counts)
save_dict('Q9 Number of custom errors thrown in methods ', throws_custom_counts)
save_dict('Q9 Number of all errors thrown in methods ', throws_counts)
save_dict('Q10 Number of generic errors caught', catches_generic_counts)
save_dict('Q10 Number of JRE errors caught', catches_jre_counts)
save_dict('Q10 Number of custom errors caught', catches_custom_counts)
save_dict('Q11 Number of "if" blocks per function', if_block_counts)
save_dict('Q11 Number of "while" blocks per function', while_block_counts)
save_dict('Q11 Number of "for" blocks per function', for_block_counts)
save_dict('Q11 Number of enhanced "for" blocks per function', enhanced_for_block_counts)
save_dict('Q11 Number of "switch" blocks per function', switch_block_counts)
save_dict('Q12 Number constants in range', {
	'Method uses constants both in and out of range +/- 32': uses_constant_in_and_out_range_count,
	'Method uses constants only in range +/- 32': uses_constant_in_range_count, 
	'Method uses constants only out of range +/- 32': uses_constant_out_range_count, 
})
save_dict('Q12 Defined final constants occurrences', defined_final_constants)
save_dict('Q12 Number value of constants being used', used_constant_numbers)

wb.save('out.xlsx')
